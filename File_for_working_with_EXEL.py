import re
import openpyxl as xl
import datetime
import time
from icecream import ic

FreeDays = []
FreeTimeForSelectedDay = []

def findTheFreeDays(month):
    FreeDays = []
    wb = xl.load_workbook('ScheduleOfWork.xlsx')
    sheet = wb[f'{month}']
    wb.active = month - 1
    #wb.active = 0
    dayNow = int(time.strftime("%d", time.localtime()))
    cols = sheet.max_column
    for i in range(dayNow + 2 - 1, cols + 1):
        cell = int(sheet.cell(row=1, column=i).value)
        if(int(cell) == 1):
            FreeDays.append(i)
    return FreeDays
def findFreeTime(day, month):
    FreeTimeForselectedDay = []
    wb = xl.load_workbook('ScheduleOfWork.xlsx')
    sheet = wb[f'{month}']
    wb.active = month - 1
    #wb.active = 0
    cols = sheet.max_row
    for i in range(3, cols + 1):
        cell = sheet.cell(row=i, column=day).value
        #ic(cell)
        if (cell == None):
            timeCell = str(sheet.cell(row = i, column = 1).value)
            FreeTimeForselectedDay.append(timeCell)
    return FreeTimeForselectedDay
def scheduleVisit(day, freeRow, month, data):
    wb = xl.load_workbook('ScheduleOfWork.xlsx')
    sheet = wb[f'{month}']
    wb.active = month - 1
    #wb.active = 0
    for i in range(2):
        sheet.cell(row = freeRow+i, column = day).value = data
    sheet.cell(row = 2, column = day).value -= 2
    wb.save(filename='ScheduleOfWork.xlsx')
def monthNowTR():
    listOfMonths = {1:"Января", 2:"Февраль", 3:"Март", 4:"Апрель",
                    5:"Май", 6:"Июнь", 7:"Июль", 8:"Август",
                    9:"Сентябрь", 10:"Октябрь", 11:"Ноябрь", 12:"Декабрь"}

    numberOfMonthNow = int(time.strftime("%m", time.localtime()))

    numberOfNextMonth = numberOfMonthNow + 1
    arrayMonthNowAndNext = [numberOfMonthNow, listOfMonths[numberOfMonthNow], numberOfNextMonth, listOfMonths[numberOfNextMonth]]


    return arrayMonthNowAndNext


def provideTheListOfProcedures():
    wb = xl.load_workbook('Price.xlsx')
    sheet = wb['price']
    wb.active = 0
    procedures = []
    numberOfProcedures = sheet.max_row
    for i in range(1, numberOfProcedures+1):
        procedure = sheet.cell(row = i, column =1).value
        procedures.append([procedure, i])
    return procedures
def provideTheCost(procedureNumber):
    wb = xl.load_workbook('Price.xlsx')
    sheet = wb['price']
    wb.active = 0
    cost = sheet.cell(row = int(procedureNumber), column = 2).value
    return cost
def getProcedure(procedureNumber):
    wb = xl.load_workbook('Price.xlsx')
    sheet = wb['price']
    wb.active = 0
    procedure = sheet.cell(row=int(procedureNumber), column=1).value
    return procedure





def findNameAndNumber(id):
    wb = xl.load_workbook('BaseOfClients.xlsx')
    sheet = wb["Лист1"]
    wb.active = 0
    cols = sheet.max_row
    for i in range(1, cols+1):
        userID = sheet.cell(row = i, column = 1).value
        if(userID == id):
            return sheet.cell(row = i, column = 2).value
    return -1
def addName(number, id):
    nameAndNumber = number
    wb = xl.load_workbook('BaseOfClients.xlsx')
    sheet = wb["Лист1"]
    wb.active = 0
    cols = sheet.max_row
    sheet.cell(row = cols+1, column = 1).value = id
    sheet.cell(row = cols+1, column = 2).value = nameAndNumber
    wb.save(filename='BaseOfClients.xlsx')

def getTime(numberOfPosition, month):
    wb = xl.load_workbook('ScheduleOfWork.xlsx')
    sheet = wb[f'{month}']
    wb.active = month - 1
    # wb.active = 0
    time = sheet.cell(row = numberOfPosition, column = 1).value
    return time



def checkPhone(phone):
    phone = str(phone)
    pattern = re.compile(r"^38\d{10}$")
    result = pattern.findall(phone)
    if len(result)>=1: return True
    else: return False














#ic(findTheFreeDays(11))
#findFreeTime(25, 11)
#ic(scheduleVisit(25,15, 11, "soijf"))




